import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

wb = Workbook()
ws = wb.active

def load_file(path):
    with open(path, "r", encoding="utf-8") as file:
        return json.load(file)
    
def configure_spreadsheet(data):
    font_header = Font(
        name="Calibri",
        size=12,
        bold=True,
        color="FFFFFFFF"
    )

    font = Font(
        name="Calibri",
        size=11,
        bold=False,
    )

    border = Border(
        left=Side(border_style="thin", color="FF000000"),
        right=Side(border_style="thin", color="FF000000"),
        top=Side(border_style="thin", color="FF000000"),
        bottom=Side(border_style="thin", color="FF000000")
    )

    alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    header = [
        "Nome", "Sexo", "Cargo", "Data de admissão", "Data de desligamento", "Status", "Salário", "Plano de saúde",
        "Cartão alimentação", "Estado", "Pontualidade", "Trabalho em equipe", "Cumprimento de metas", "Data de avaliação"
    ]

    ws.append(header)

    for person in data:
        row = []

        for value in person.values():
            row.append(value)

        ws.append(row)
    
    for row in ws:
        for cell in row:
            cell.font = font
            cell.border = border
            cell.alignment = alignment

    for row in ws["A1:N1"]:
        for cell_header in row:
            cell_header.font = font_header
            cell_header.fill = header_fill

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[column_letter].width = max_length + 4

data = load_file("dados_gerados.json")
configure_spreadsheet(data)

wb.save("report.xlsx")